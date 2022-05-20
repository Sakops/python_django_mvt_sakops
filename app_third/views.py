from email import utils
from django.shortcuts import render, redirect
from app_third import models, utils
from django.urls import reverse
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# тут только "логика" - функции для обработки и возврат данных

class Todo: 
    def __init__(self, description, name="name1"): 
        self.description = description
        self.name = name
    def counter(self, external_value): 
        self.value += external_value
    @staticmethod 
    def count(value, extra): 
        return value + extra
obj = Todo("привет", "привет")

def index(request):
    return render(request, 'app_third/pages/index.html')


def home(request):
    return render(request, 'app_third/pages/home.html')


def about(request):
    return render(request, 'app_third/pages/about.html')

def adminpage(request):
    if request.method == "POST": 
        excel = request.FILES.get('excel', None)
    context = {

    }
    names = [
        ["sam", "robert", "alex"], 
        [1, 2, 3], 
        [6, 7, 8]
    ]
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # # Data can be assigned directly to cells
    # ws['A1'] = 42
    for i in names: 
        for j in i: 
            x = names.index(i) + 1
            y = get_column_letter(i.index(j) + 1)
            print(f"x = {x}, y = {y}", j)
            ws[f"{y}{x}"] = j
            
    # # Rows can also be appended
    # ws.append([1, 2, 3])

    # # Python types will automatically be converted
    # import datetime
    # ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")

    return render(request, 'app_third/pages/AdminPage.html', context)

def login(request): 
    return render(request, 'app_third/pages/login.html')
def origin_home(request):
    return render(request, 'app_third/pages/origin_home.html')


def todo_detail(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)
    context = {
        "todo": obj
    }
    return render(request, 'app_third/pages/DetailTodo.html', context)


def todo_list(request):
    page_obj = utils.CustomPaginator.get_page(
        objs=models.Task.objects.all(),
        limit=2,
        current_page=request.GET.get('page')
    )
    context = {"list": None, "page": page_obj}
    return render(request, 'app_third/pages/todo_list.html', context)

def todo_create(request):
    if request.method == "POST":
        title1 = request.POST.get("title", "заголовок по умолчанию")
        description1 = request.POST.get("description", "описание по умолчанию")
        obj = models.Task.objects.create(
            title=title1,
            description=description1
        )
        obj.save()
    context = {}
    return render(request, 'app_third/pages/CreateTodo.html', context)
def todo_delete(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)
    obj.delete()
    return redirect(reverse('todo_list', args=()))
def todo_update(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)
    # obj.is_completed = not obj.is_completed
    if obj.is_completed:
        obj.is_completed = False
    else:
        obj.is_completed = True
    obj.save()
    return redirect(reverse('todo_list', args=()))

def todo_change_data(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)
    if request.method == "POST":
        title1 = request.POST.get("title", "заголовок по умолчанию")
        description1 = request.POST.get("description", "описание по умолчанию")
        
        obj = models.Task.objects.get(id=todo_id)

        obj.title = title1
        obj.description = description1
        obj.save()


    context = {
        "todo": obj
    }
    return render(request, 'app_third/pages/ChangeTodo.html', context)
